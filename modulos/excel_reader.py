"""
excel_reader.py
----------------
Lectura del Excel de personas y generacion de la plantilla Excel estandar.

Columnas esperadas (en este orden, encabezados exactos):
    DNI, Nombres, Apellido paterno, Apellido materno, Fecha de nacimiento,
    Numero de carnet, Fecha de emision, Fecha de vencimiento, Empresa, Puesto,
    Fotografia

Reglas de negocio (ver validations.py para el detalle):
    - DNI, Nombres, Apellidos y Fecha de nacimiento y Fecha de emision son
      obligatorios.
    - Numero de carnet y Fecha de vencimiento pueden llegar vacios: se
      autogeneran/calculan segun la configuracion (ver modules/config.py).
    - Empresa, Puesto y Fotografia son opcionales.
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Encabezados oficiales del Excel (orden fijo, no modificar sin actualizar
# validations.py y app.py).
COLUMNS: List[str] = [
    "DNI",
    "Nombres",
    "Apellido paterno",
    "Apellido materno",
    "Fecha de nacimiento",
    "Numero de carnet",
    "Fecha de emision",
    "Fecha de vencimiento",
    "Empresa",
    "Puesto",
    "Fotografia",
]

REQUIRED_COLUMNS: List[str] = [
    "DNI",
    "Nombres",
    "Apellido paterno",
    "Apellido materno",
    "Fecha de nacimiento",
    "Fecha de emision",
]

MAX_EXCEL_SIZE_MB = 10
ALLOWED_EXCEL_EXTENSIONS = {".xlsx"}


class ExcelFormatError(Exception):
    """Se lanza cuando el archivo no tiene el formato/columnas esperadas."""


@dataclass
class ExcelLoadResult:
    dataframe: pd.DataFrame
    missing_columns: List[str]
    extra_columns: List[str]


def build_template_workbook() -> Workbook:
    """Construye el libro de Excel estandar descargable (con datos de ejemplo)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Personas"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    # Fila de ejemplo (datos ficticios, no reales) para orientar al usuario.
    ejemplo = [
        "12345678",
        "Maria Fernanda",
        "Quispe",
        "Rojas",
        "1995-04-12",
        "",  # numero de carnet -> puede autogenerarse
        "2026-01-15",
        "",  # fecha de vencimiento -> puede autocalcularse
        "Empresa Ejemplo S.A.C.",
        "Auxiliar de cocina",
        "maria_fernanda.jpg",
    ]
    for col_idx, value in enumerate(ejemplo, start=1):
        ws.cell(row=2, column=col_idx, value=value)

    ws.freeze_panes = "A2"

    # Validacion simple de fechas como texto guia (comentario en la hoja).
    ws["A1"].comment = None

    # Hoja de instrucciones.
    ws2 = wb.create_sheet("Instrucciones")
    instrucciones = [
        "Instrucciones para completar el Excel de carnets de sanidad",
        "",
        "1. No cambie los encabezados de la hoja 'Personas'.",
        "2. DNI, Nombres, Apellido paterno, Apellido materno, Fecha de nacimiento",
        "   y Fecha de emision son obligatorios.",
        "3. Formato de fechas sugerido: AAAA-MM-DD (ej. 2026-01-15).",
        "4. Numero de carnet puede dejarse vacio si la aplicacion lo autogenera.",
        "5. Fecha de vencimiento puede dejarse vacia si se calcula automaticamente",
        "   segun el periodo de vigencia configurado.",
        "6. Empresa y Puesto son opcionales.",
        "7. Fotografia: escriba el nombre exacto del archivo de imagen que",
        "   subira junto con el Excel (ej. juan_perez.jpg). Es opcional.",
        "8. No agregue columnas de Resultado, Observaciones ni Codigo QR.",
    ]
    for i, line in enumerate(instrucciones, start=1):
        ws2.cell(row=i, column=1, value=line)
    ws2.column_dimensions["A"].width = 80
    ws2["A1"].font = Font(bold=True, size=12)

    return wb


def template_bytes() -> bytes:
    """Devuelve los bytes .xlsx de la plantilla estandar (para descarga)."""
    wb = build_template_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def validate_extension(filename: str) -> None:
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXCEL_EXTENSIONS:
        raise ExcelFormatError(
            f"Formato de archivo no permitido: '{ext}'. Solo se acepta .xlsx"
        )


def load_excel(file_like, filename: str = "archivo.xlsx", max_size_mb: int = MAX_EXCEL_SIZE_MB) -> ExcelLoadResult:
    """Carga y normaliza el Excel de personas.

    Parameters
    ----------
    file_like: objeto tipo archivo (BytesIO, UploadedFile de Streamlit, o ruta)
    filename: nombre original, usado para validar extension
    max_size_mb: limite de tamano en MB

    Returns
    -------
    ExcelLoadResult con el DataFrame normalizado (columnas como texto/str,
    salvo fechas que se intentan parsear en validations.py) y las columnas
    faltantes/sobrantes detectadas.
    """
    validate_extension(filename)

    # Tamano del archivo
    data = file_like.read() if hasattr(file_like, "read") else Path(file_like).read_bytes()
    size_mb = len(data) / (1024 * 1024)
    if size_mb > max_size_mb:
        raise ExcelFormatError(
            f"El archivo pesa {size_mb:.1f} MB y excede el limite permitido de {max_size_mb} MB."
        )

    try:
        df = pd.read_excel(io.BytesIO(data), sheet_name="Personas", dtype=str)
    except ValueError:
        # Si no existe la hoja "Personas", se usa la primera hoja disponible.
        try:
            df = pd.read_excel(io.BytesIO(data), sheet_name=0, dtype=str)
        except Exception as exc:  # noqa: BLE001
            raise ExcelFormatError(f"No se pudo leer el archivo Excel: {exc}") from exc
    except Exception as exc:  # noqa: BLE001
        raise ExcelFormatError(f"No se pudo leer el archivo Excel: {exc}") from exc

    df.columns = [str(c).strip() for c in df.columns]

    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    extra = [c for c in df.columns if c not in COLUMNS]

    # Asegura que existan todas las columnas del esquema (las que falten se
    # crean vacias para no romper el resto del flujo; la ausencia de
    # obligatorias se reporta igualmente en `missing`).
    for c in COLUMNS:
        if c not in df.columns:
            df[c] = ""

    df = df[COLUMNS].fillna("")
    df = df.reset_index(drop=True)
    df.insert(0, "_fila_excel", df.index + 2)  # fila real en Excel (encabezado = fila 1)

    return ExcelLoadResult(dataframe=df, missing_columns=missing, extra_columns=extra)
